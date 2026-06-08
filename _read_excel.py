import openpyxl
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def read_xlsx(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    result = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            if all(cell is None for cell in row):
                continue
            cleaned = []
            for cell in row:
                if cell is None:
                    cleaned.append("")
                elif isinstance(cell, (int, float)):
                    cleaned.append(str(cell))
                else:
                    try:
                        cleaned.append(str(cell))
                    except Exception:
                        cleaned.append(repr(cell))
            rows.append(cleaned)
        result[sheet_name] = rows
    return result

files = [
    ("Linkedin Crawl Report", "D:/CrawlDataLinkedin/Linkedin Crawl Report.xlsx"),
    ("nguyenggsheet", "D:/CrawlDataLinkedin/nguyenggsheet.xlsx"),
]

for name, fp in files:
    print(f"\n{'='*60}")
    print(f"FILE: {name}")
    print(f"{'='*60}")
    try:
        data = read_xlsx(fp)
        print(f"Sheets: {list(data.keys())}")
        for sheet, rows in data.items():
            print(f"\n--- Sheet: '{sheet}' ({len(rows)} rows) ---")
            for i, row in enumerate(rows[:50]):
                print(f"  [{i}] {row}")
            if len(rows) > 50:
                print(f"  ... ({len(rows) - 50} more rows)")
    except Exception as e:
        import traceback
        traceback.print_exc()
